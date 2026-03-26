import os
import sacrebleu
from tqdm import tqdm
from ontology_manager import OntologyManager

from translator import OntologyAwareTranslator

# Configuration
VERBOSE = False

def main():
    print("Initializing NMT System with Ontology Integration...")

    ontology_manager = OntologyManager()
    base_translator = OntologyAwareTranslator(verbose=VERBOSE, load_adapters=False)
    ft_translator = OntologyAwareTranslator(verbose=VERBOSE, load_adapters=True)

    import openpyxl
    
    test_data = []
    try:
        # Load English source
        wb_en = openpyxl.load_workbook("data/en.xlsx")
        sheet_en = wb_en.active
        
        # Load French reference
        wb_fr = openpyxl.load_workbook("data/fr.xlsx")
        sheet_fr = wb_fr.active
        max_rows = min(sheet_en.max_row, sheet_fr.max_row)
        
        print(f"Loading test data from Excel (found {max_rows-1} rows)...")
        
        for i in range(2, max_rows + 1):
            source_text = sheet_en.cell(row=i, column=1).value
            reference_text = sheet_fr.cell(row=i, column=1).value
            
            if source_text and reference_text:
                test_data.append({
                    "source": str(source_text).strip(),
                    "reference": str(reference_text).strip()
                })
                
    except Exception as e:
        print(f"Error loading Excel files: {e}")
        print("Falling back to hardcoded test data.")
        exit(1)
    
    print("\n" + "="*50)
    print("STARTING EVALUATION")
    print("="*50 + "\n")

    baseline_hyps = []
    constrained_hyps = []
    references = []

    for i, item in enumerate(tqdm(test_data, desc="Translating")):
        sentence = item["source"]
        reference = item["reference"]
        references.append(reference)

        # 1. Baseline Translation (Base Model, No Constraints)
        baseline_translation = base_translator.translate(sentence)
        baseline_hyps.append(baseline_translation)
        
        # 2. Ontology Lookup
        concepts = ontology_manager.identify_concepts(sentence)
        constraints = []
        if concepts:
            for concept in concepts:
                if concept['context_score'] > 0:
                     constraints.append(concept)

        if constraints:
            constrained_translation = ft_translator.translate(sentence, constraints=constraints)
        else:
            constrained_translation = ft_translator.translate(sentence)
        
        constrained_hyps.append(constrained_translation)

    # Calculate chrF++ Scores
    print("\n" + "="*50)
    print("EVALUATION METRICS")
    print("="*50)
    
    # 1. chrF++
    baseline_chrf = sacrebleu.corpus_chrf(baseline_hyps, [references], word_order=2)
    constrained_chrf = sacrebleu.corpus_chrf(constrained_hyps, [references], word_order=2)
    
    print(f"Baseline chrF++:    {baseline_chrf.score:.2f}")
    print(f"Constrained chrF++: {constrained_chrf.score:.2f}")
    
    # 2. BLEU (Standard 13a tokenizer)
    baseline_bleu = sacrebleu.corpus_bleu(baseline_hyps, [references], force=True)
    constrained_bleu = sacrebleu.corpus_bleu(constrained_hyps, [references], force=True)

    print(f"Baseline BLEU:      {baseline_bleu.score:.2f}")
    print(f"Constrained BLEU:   {constrained_bleu.score:.2f}")

    # 3. spBLEU (BLEU on SentencePiece tokens)
    # We use the model's tokenizer to tokenize hyps and refs, then compute BLEU on the tokens
    def tokenize_for_spbleu(texts, tokenizer):
        tokenized = []
        for text in texts:
            # Tokenize and join with spaces
            tokens = tokenizer.tokenize(text)
            tokenized.append(" ".join(tokens))
        return tokenized

    baseline_hyps_sp = tokenize_for_spbleu(baseline_hyps, base_translator.tokenizer)
    constrained_hyps_sp = tokenize_for_spbleu(constrained_hyps, ft_translator.tokenizer)
    references_sp = tokenize_for_spbleu(references, base_translator.tokenizer)

    # Use tokenize='none' because we already tokenized
    baseline_spbleu = sacrebleu.corpus_bleu(baseline_hyps_sp, [references_sp], tokenize='none')
    constrained_spbleu = sacrebleu.corpus_bleu(constrained_hyps_sp, [references_sp], tokenize='none')

    print(f"Baseline spBLEU:    {baseline_spbleu.score:.2f}")
    print(f"Constrained spBLEU: {constrained_spbleu.score:.2f}")

    # Save results to output file
    import random
    from datetime import datetime

    output_file = "results_output.txt"
    print(f"\nSaving results to {output_file}...")

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("="*70 + "\n")
        f.write("NMT SYSTEM WITH ONTOLOGY INTEGRATION - EVALUATION RESULTS\n")
        f.write("="*70 + "\n\n")

        # Model Information
        f.write("MODEL INFORMATION\n")
        f.write("-"*70 + "\n")
        f.write(f"Base Model: {base_translator.model.config._name_or_path}\n")
        f.write(f"Fine-Tuned Model: {ft_translator.model.config._name_or_path} (with adapters)\n")
        f.write(f"Device: {base_translator.device}\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Test Set Size: {len(test_data)} sentences\n")
        f.write("\n")

        # Evaluation Results
        f.write("EVALUATION METRICS\n")
        f.write("-"*70 + "\n")
        f.write(f"{'Metric':<25} {'Baseline':<15} {'Constrained':<15} {'Δ':<10}\n")
        f.write("-"*70 + "\n")

        chrf_delta = constrained_chrf.score - baseline_chrf.score
        bleu_delta = constrained_bleu.score - baseline_bleu.score
        spbleu_delta = constrained_spbleu.score - baseline_spbleu.score

        f.write(f"{'chrF++':<25} {baseline_chrf.score:<15.2f} {constrained_chrf.score:<15.2f} {chrf_delta:+.2f}\n")
        f.write(f"{'BLEU':<25} {baseline_bleu.score:<15.2f} {constrained_bleu.score:<15.2f} {bleu_delta:+.2f}\n")
        f.write(f"{'spBLEU':<25} {baseline_spbleu.score:<15.2f} {constrained_spbleu.score:<15.2f} {spbleu_delta:+.2f}\n")
        f.write("\n")

        # Random Sample Translations
        f.write("SAMPLE TRANSLATIONS (10 random examples)\n")
        f.write("="*70 + "\n\n")

        # Select random samples
        sample_indices = random.sample(range(len(test_data)), min(10, len(test_data)))

        for idx, i in enumerate(sample_indices, 1):
            f.write(f"Sample {idx} - Question {i}\n")
            f.write("-"*70 + "\n")
            f.write(f"Source:      {test_data[i]['source']}\n")
            f.write(f"Reference:   {references[i]}\n")
            f.write(f"Baseline:    {baseline_hyps[i]}\n")
            f.write(f"Constrained: {constrained_hyps[i]}\n")

            # Check if ontology was used for this example
            concepts = ontology_manager.identify_concepts(test_data[i]['source'])
            used_constraints = [c['term'] for c in concepts if c['context_score'] > 0]
            if used_constraints:
                f.write(f"Constraints: {', '.join(used_constraints)}\n")
            
            f.write("\n")

        f.write("="*70 + "\n")
        f.write("END OF REPORT\n")
        f.write("="*70 + "\n")

    print(f"✓ Results saved to {output_file}")

if __name__ == "__main__":
    main()